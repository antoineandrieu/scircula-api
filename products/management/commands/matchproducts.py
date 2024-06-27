#!/usr/bin/env python3

import argparse
import difflib
import os

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook, load_workbook
from products.models import Product


# Convert boolean arguments to Python boolean
def str2bool(v):
    if isinstance(v, bool):
        return v
    if v.lower() in ("yes", "true", "t", "y", "1"):
        return True
    elif v.lower() in ("no", "false", "f", "n", "0"):
        return False
    else:
        raise argparse.ArgumentTypeError("Boolean value expected.")


class Command(BaseCommand):
    help = "Add vendor to product from a spreadsheet"

    def add_arguments(self, parser):
        parser.add_argument(
            "-i", "--input", required=True, help="Product Listing File Path"
        )
        parser.add_argument(
            "-o", "--output", required=True, help="Updated Product Listing File Path"
        )

    def handle(self, *args, **options):
        input_wb = load_workbook(options["input"])
        input_ws = input_wb.worksheets[0]

        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.cell(row=1, column=1).value = "Brand"
        output_ws.cell(row=1, column=2).value = "External Id"
        output_ws.cell(row=1, column=3).value = "Vendor Product Name"
        output_ws.cell(row=1, column=4).value = "SCIRCULA Product Name"
        output_ws.cell(row=1, column=5).value = "Match Ratio"
        output_ws.cell(row=1, column=6).value = "SCIRCULA Id"
        output_ws.cell(row=1, column=7).value = "Sizes"
        output_ws.cell(row=1, column=8).value = "Notes"
        output_row = 2

        new_products = {}
        for in_row_idx in range(2, input_ws.max_row):
            name = input_ws.cell(row=in_row_idx, column=1).value
            external_id = input_ws.cell(row=in_row_idx, column=2).value
            brand = input_ws.cell(row=in_row_idx, column=3).value
            sizes = input_ws.cell(row=in_row_idx, column=7).value
            if not new_products.get(brand):
                new_products[brand] = []

            new_products[brand].append(
                {"name": name, "external_id": external_id, "sizes": sizes}
            )

        out_row_idx = 2
        for brand, products_list in new_products.items():
            output_ws.cell(row=out_row_idx, column=1).value = brand
            products = list(
                Product.objects.filter(brand__iexact=brand).values("name", "id", "code")
            )
            self.stdout.write(f"File product count for {brand}: {len(products_list)}")
            self.stdout.write(f"DB product count for {brand}: {len(products)}")
            for new_product in products_list:
                new_product_name = new_product["name"]
                output_ws.cell(row=out_row_idx, column=2).value = new_product[
                    "external_id"
                ]
                output_ws.cell(row=out_row_idx, column=3).value = new_product_name
                match_flag = False
                for product in products:
                    product_name = product["name"]
                    if not product_name:
                        product_name = product["code"] or "no name"
                    diff_ratio = difflib.SequenceMatcher(
                        None, new_product_name, product_name
                    ).ratio()
                    if diff_ratio > 0.2:
                        match_flag = True
                        output_ws.cell(row=out_row_idx, column=4).value = product_name
                        output_ws.cell(row=out_row_idx, column=5).value = round(
                            diff_ratio, 3
                        )
                        output_ws.cell(row=out_row_idx, column=6).value = product["id"]
                        output_ws.cell(row=out_row_idx, column=7).value = new_product[
                            "sizes"
                        ]
                        out_row_idx += 1
                if match_flag == False:
                    out_row_idx += 1
        output_wb.save(options["output"])
