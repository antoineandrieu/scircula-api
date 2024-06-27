#!/usr/bin/env python3

import argparse
import difflib
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError
from openpyxl import Workbook, load_workbook
from psycopg2.errors import UniqueViolation
from products.models import Product, ProductVendor, ProductSize, ProductSizeVendor, Size
from vendors.models import Vendor


class Command(BaseCommand):
    help = "Add vendor to product from a spreadsheet"

    ws_row = 2
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.cell(row=1, column=1).value = "Brand"
    output_ws.cell(row=1, column=2).value = "Product"
    output_ws.cell(row=1, column=3).value = "Product Size"

    def add_arguments(self, parser):
        parser.add_argument(
            "-i", "--input", required=True, help="Product Listing File Path"
        )
        parser.add_argument("-ve", "--vendor", required=True, help="Vendor URL")

    def parse_sizes_vendor(self, product, vendor, raw_sizes):
        raw_sizes = raw_sizes.replace("\n", "")
        raw_sizes = raw_sizes.replace("\t", "")

        sizes = raw_sizes.split(",")
        sizes_data = []
        for raw_size in sizes:
            external_id = None
            if ":" in raw_size:
                split_raw_size = raw_size.split(":")
                name = split_raw_size[0]
                external_id = split_raw_size[1].strip()
            else:
                name = raw_size.rstrip()

            if all(symbol in name for symbol in ["W", "L"]):
                raw_name = name.translate({ord(char): None for char in "WL "})
                waist_size = raw_name[:2]
                sizes_data.append(
                    {"name": waist_size, "inseam": False, "composed": False}
                )
                length_size = raw_name[-2:]
                sizes_data.append(
                    {"name": length_size, "inseam": True, "composed": False}
                )
                composed_size = f"{waist_size} / {length_size}"

                sizes_data.append(
                    {
                        "name": composed_size,
                        "vendor_name": name.rstrip(),
                        "external_id": external_id,
                        "inseam": False,
                        "composed": True,
                    }
                )
            # elif "/" in name:
            #     raw_name = name.translate({ord(char): None for char in "/ "})
            #     waist_size = raw_name[:2]
            #     sizes_data.append(
            #         {"name": waist_size, "inseam": False, "composed": False}
            #     )
            #     length_size = raw_name[-2:]
            #     sizes_data.append(
            #         {"name": length_size, "inseam": True, "composed": False}
            #     )
            #     composed_size = f"{waist_size} / {length_size}"

            #     sizes_data.append(
            #         {
            #             "name": composed_size,
            #             "vendor_name": name.rstrip(),
            #             "external_id": external_id,
            #             "inseam": False,
            #             "composed": True,
            #         }
            #     )
            elif "W" in name and not "L" in name:
                sizes_data.append(
                    {
                        "name": name.lower(),
                        # "name": name.rstrip()[1:],
                        "vendor_name": name.rstrip(),
                        "external_id": external_id,
                        "inseam": False,
                        "composed": False,
                    }
                )

            else:
                sizes_data.append(
                    {
                        "name": name.rstrip(),
                        "vendor_name": name.rstrip(),
                        "external_id": external_id,
                        "inseam": False,
                        "composed": False,
                    }
                )

        return sizes_data

    def create_sizes_vendor(self, product, vendor, raw_sizes):
        sizes_data = self.parse_sizes_vendor(product, vendor, raw_sizes)

        for size in sizes_data:
            # Create size & product size if doesn't exist
            db_size = None
            try:
                db_size = Size.objects.get(name=size["name"])
            except Size.DoesNotExist:
                self.output_ws.cell(row=self.ws_row, column=1).value = product.brand
                self.output_ws.cell(row=self.ws_row, column=2).value = product.name
                self.output_ws.cell(row=self.ws_row, column=3).value = size["name"]
                self.ws_row += 1
                self.stderr.write(f"Size {size['name']} does not exists")

            if db_size:
                product_size = None
                try:
                    product_size = ProductSize.objects.get(
                        product=product,
                        size=db_size,
                        inseam=size["inseam"],
                        composed=size["composed"],
                    )
                except ProductSize.DoesNotExist:
                    self.output_ws.cell(row=self.ws_row, column=1).value = product.brand
                    self.output_ws.cell(row=self.ws_row, column=2).value = product.name
                    self.output_ws.cell(row=self.ws_row, column=3).value = db_size.name
                    self.ws_row += 1
                    self.stderr.write(
                        f"Product Size {size['name']} for product {product.code} does not exists"
                    )
                if product_size:
                    try:
                        ProductSizeVendor.objects.get(
                            product_size=product_size,
                            vendor=vendor,
                        )
                    except ProductSizeVendor.DoesNotExist:
                        try:
                            ProductSizeVendor.objects.create(
                                product_size=product_size,
                                vendor=vendor,
                                external_id=size.get("external_id"),
                                name=size["name"],
                            )
                        except IntegrityError:
                            self.stderr.write(
                                f"Product Size {size['name']} relation with vendor {vendor} already exists"
                            )

    def handle(self, *args, **options):
        vendor_url = options["vendor"]
        try:
            vendor = Vendor.objects.get(url=vendor_url)
        except Vendor.DoesNotExist:
            raise CommandError(f"Vendor with URL {options['vendor']} does not exist")

        input_wb = load_workbook(options["input"])
        input_ws = input_wb.worksheets[0]

        for in_row_idx in range(2, input_ws.max_row):
            external_id = input_ws.cell(row=in_row_idx, column=2).value
            name = input_ws.cell(row=in_row_idx, column=3).value
            code = input_ws.cell(row=in_row_idx, column=4).value
            scircula_id = input_ws.cell(row=in_row_idx, column=6).value
            raw_sizes = input_ws.cell(row=in_row_idx, column=7).value
            product = None
            if scircula_id:
                try:
                    product = Product.objects.get(code=code)
                except Product.DoesNotExist:
                    self.stderr.write(f"Product {name} with code {code} does not exist")

                if product and name:
                    try:
                        productvendor = ProductVendor.objects.get(
                            product=product,
                            vendor=vendor,
                        )
                        try:
                            productvendor.external_id = external_id
                            productvendor.save()
                        except (IntegrityError, UniqueViolation):
                            self.stderr.write(
                                f"Product {name} relation with vendor {vendor} already exists with external id {external_id}"
                            )

                    except ProductVendor.DoesNotExist:
                        try:
                            productvendor = ProductVendor.objects.create(
                                product=product,
                                vendor=vendor,
                                name=name,
                                external_id=external_id,
                            )
                            self.stdout.write(
                                f"Product {name} relation with vendor {vendor} created with id {productvendor.id}"
                            )
                        except IntegrityError:
                            self.stderr.write(
                                f"Product {name} relation with vendor {vendor} already exists with external id {external_id}"
                            )

                    self.create_sizes_vendor(product, vendor, raw_sizes)
        self.output_wb.save("missing-products.xlsx")
