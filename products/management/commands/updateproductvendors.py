#!/usr/bin/env python3

import argparse
import difflib
import os

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook, load_workbook
from products.models import ProductVendor


# Convert boolean arguments to Python boolean
def str2bool(v):
    if isinstance(v, bool):
        return v
    if v.lower() in ("yes", "true", "t", "y", "1"):
        return True
    elif v.lower() in ("no", "false", "f", "n", "0"):
        return False
    else:
        raise argparse.ArgumentTypeError("Boolean value expected.")


class Command(BaseCommand):
    help = "Add vendor to product from a spreadsheet"

    def add_arguments(self, parser):
        parser.add_argument(
            "-i", "--input", required=True, help="Product Listing File Path"
        )

    def handle(self, *args, **options):
        input_wb = load_workbook(options["input"])
        input_ws = input_wb.worksheets[0]

        new_products = {}
        for in_row_idx in range(2, input_ws.max_row):
            name = input_ws.cell(row=in_row_idx, column=1).value
            external_id = input_ws.cell(row=in_row_idx, column=2).value
            brand = input_ws.cell(row=in_row_idx, column=3).value
            sizes = input_ws.cell(row=in_row_idx, column=7).value
            if not new_products.get(brand):
                new_products[brand] = []

            new_products[brand].append(
                {"name": name, "external_id": external_id, "sizes": sizes}
            )

        out_row_idx = 2
        for brand, products_list in new_products.items():
            for new_product in products_list:
                new_product_name = new_product["name"]
                if new_product_name.endswith("-1"):
                    new_product_name = new_product_name[:-2]
                    self.stdout.write(f"Product Name: {new_product_name}")
                try:
                    old_product = ProductVendor.objects.get(
                        product__code=new_product_name,
                        vendor__url="https://kuyichi.com",
                    )
                except ProductVendor.DoesNotExist:
                    try:
                        old_products = ProductVendor.objects.filter(
                            product__code__contains=new_product_name,
                            vendor__url="https://kuyichi.com",
                        )
                        if len(old_products) == 1:
                            old_product = old_products[0]
                        else:
                            print(old_products)
                    except ProductVendor.DoesNotExist:
                        self.stderr.write(
                            f"Can't find product vendor: {new_product_name}"
                        )
