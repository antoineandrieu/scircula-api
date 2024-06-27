#!/usr/bin/env python3

import argparse
import difflib
import os

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook, load_workbook
from products.models import Product, ProductAttribute, ProductSize


class Command(BaseCommand):
    help = "Extract product data write it to a file"

    dataset_sizes = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"]
    # dataset_sizes = [
    #     "24",
    #     "25",
    #     "26",
    #     "27",
    #     "28",
    #     "29",
    #     "30",
    #     "31",
    #     "32",
    #     "33",
    #     "34",
    #     "35",
    #     "36",
    #     "37",
    #     "38",
    #     "40",
    #     "42",
    #     "44",
    #     "46",
    #     "48",
    # ]
    # size_conversion = {"Small": "S", "Medium": "M", "Large": "L", "X Large": "XL"}
    # size_conversion = {
    #     "w25": "25",
    #     "w26": "26",
    #     "w27": "27",
    #     "w28": "28",
    #     "w29": "29",
    #     "w30": "30",
    #     "w31": "31",
    #     "W25": "25",
    #     "W26": "26",
    #     "W27": "27",
    #     "W28": "28",
    #     "W29": "29",
    #     "W30": "30",
    #     "W31": "31",
    # }
    size_conversion = {"Small": "S", "Medium": "M", "Large": "L", "X Large": "XL"}
    column_offset = 7
    missing_sizes = set()
    missing_products = set()
    worksheet = None
    excluded_brands = ["scircula", "pug parties"]

    def add_arguments(self, parser):
        parser.add_argument("-o", "--output", help="Updated Product Listing File Path")

    def write_headers(self):
        self.worksheet.cell(row=1, column=1).value = "Id"
        self.worksheet.cell(row=1, column=2).value = "Code"
        self.worksheet.cell(row=1, column=3).value = "Brand"
        self.worksheet.cell(row=1, column=4).value = "Category"
        self.worksheet.cell(row=1, column=5).value = "Gender"
        self.worksheet.cell(row=1, column=6).value = "Measurement Point"

        for col_no in range(len(self.dataset_sizes)):
            self.worksheet.cell(
                row=1, column=col_no + self.column_offset
            ).value = self.dataset_sizes[col_no]

    def get_sizes_specs(self, product):
        sizes_specs = {}
        attributes = ProductAttribute.objects.filter(product_size__product=product)
        for attr in attributes:
            tag = attr.tag
            value = attr.value
            size = attr.product_size.size.name
            if sizes_specs.get(tag):
                if not sizes_specs[tag].get(size):
                    sizes_specs[tag][size] = value
            else:
                sizes_specs[tag] = {size: value}

        return sizes_specs

    def get_products_specs(self, products):
        products_specs = {}
        for product in products:
            sizes = self.get_sizes_specs(product)
            products_specs[product.code] = {
                "id": product.id,
                "brand": product.brand,
                "category": product.category.name,
                "gender": product.gender,
                "sizes": self.get_sizes_specs(product),
            }

        return products_specs

    def convert_size(self, raw_size):
        return self.size_conversion.get(raw_size, raw_size.upper())

    def write_measurements(self, row_no, raw_size, value, product):
        write_flag = False
        size = self.convert_size(raw_size)
        try:
            column_no = self.dataset_sizes.index(size) + self.column_offset
            self.worksheet.cell(row=row_no, column=column_no).value = value
            write_flag = True
        except ValueError:
            self.missing_sizes.add(raw_size)
            self.missing_products.add(product)

        return write_flag

    def write_specs(self, row_no, code, specs, tag):
        self.worksheet.cell(row=row_no, column=1).value = specs["id"]
        self.worksheet.cell(row=row_no, column=2).value = code
        self.worksheet.cell(row=row_no, column=3).value = specs["brand"]
        self.worksheet.cell(row=row_no, column=4).value = specs["category"]
        self.worksheet.cell(row=row_no, column=5).value = specs["gender"]
        self.worksheet.cell(row=row_no, column=6).value = tag

    def write_products_specs(self, products_specs):
        row_no = 2
        for code, specs in products_specs.items():
            if specs["brand"] not in self.excluded_brands and (
                specs["category"] == "pants" or specs["category"] == "kuyichi_pants"
            ):
                # if specs["brand"] not in self.excluded_brands:
                for tag, sizes in specs["sizes"].items():
                    write_flag = False
                    for size, value in sizes.items():
                        write_flag = self.write_measurements(row_no, size, value, code)
                    if write_flag:
                        self.write_specs(row_no, code, specs, tag)
                        row_no = row_no + 1

    def handle(self, *args, **options):
        output_wb = Workbook()
        self.worksheet = output_wb.active
        self.write_headers()
        # products = Product.objects.exclude(category__name="pants")
        products = Product.objects.filter(category__name="pants")
        products_specs = self.get_products_specs(products)
        self.write_products_specs(products_specs)

        print(f"Found {products.count()} products")
        print(f"Missing products: {list(sorted(self.missing_products))}")
        print(f"Missing sizes: {list(sorted(self.missing_sizes))}")

        output_wb.save(options["output"] or "/opt/app/output.xlsx")
